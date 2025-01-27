import io
import json
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

def get_market_rate(part):
    response = requests.post('https://mosapi.sinewave.co.in/stocks/', json={'tickers': part}, verify=False)
    string = response.json()
    json_res = json.loads(string)
    if len(json_res) > 0:
        data = json_res[0]
        return data
    else:
        return None

def get_market_rate_value(part):
    data = get_market_rate(part)
    if data:
        return data['Adj Close']
    else:
        return None

def get_strategy(parts, days):
    response = requests.post('http://filerenderapi.sinewave.co.in/api/', json={'tickers': parts, 'per': int(days)}, verify=False)
    string = response.content
    toread = io.BytesIO(string)
    toread.seek(0)
    return toread

def get_strategy_values(parts, days):
    toread = get_strategy(parts, days)
    result = []
    wb = load_workbook(toread, read_only=True)
    for part in parts:
        if part not in wb.sheetnames:
            continue
        df = pd.read_excel(toread, sheet_name=part)
        last = df.tail(1).iloc[0]
        record = {
            'part': part,
            'date': last['Date'],
            'trigger': last['ATR']
        }
        result.append(record)
    return result

def get_strategy_file(parts, days):
    bytes_data = get_strategy(parts, days)
    wb = load_workbook(io.BytesIO(bytes_data))
    virtual_workbook = io.BytesIO()
    with ExcelWriter(virtual_workbook) as ew:
        ew.write_wb(wb)
    return virtual_workbook.getvalue()
